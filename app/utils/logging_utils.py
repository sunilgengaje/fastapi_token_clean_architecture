# app/utils/logging_utils.py
from pathlib import Path
from datetime import datetime
from typing import List
import os

# Optional Excel support
try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# FastAPI Request import for helper
from fastapi import Request

# Paths
PROJECT_ROOT = Path.cwd()
LOG_BASE = PROJECT_ROOT / "logs"
SESSION_DIR = LOG_BASE / "sessions"
LOG_BASE.mkdir(parents=True, exist_ok=True)
SESSION_DIR.mkdir(parents=True, exist_ok=True)

# Rotating log settings
MAX_FILE_SIZE = 2 * 1024 * 1024  # 2MB
LOG_FILENAME_PREFIX = "activity_log_"
LOG_EXT = "txt"

def get_client_ip(request: Request) -> str:
    """
    Extract client IP from headers or request.client.
    """
    xff = request.headers.get("x-forwarded-for")
    if xff:
        return xff.split(",")[0].strip()
    client = getattr(request, "client", None)
    if client:
        return client.host
    return "unknown"

def _active_log_index() -> int:
    """
    Find the current active rotating file index, creating files as needed.
    """
    idx = 1
    while True:
        p = LOG_BASE / f"{LOG_FILENAME_PREFIX}{idx}.{LOG_EXT}"
        if not p.exists():
            return idx
        if p.stat().st_size < MAX_FILE_SIZE:
            return idx
        idx += 1

def get_active_log_file() -> Path:
    idx = _active_log_index()
    return LOG_BASE / f"{LOG_FILENAME_PREFIX}{idx}.{LOG_EXT}"

def append_rotating_log(
    method: str,
    path: str,
    status: int,
    ip: str,
    ua: str,
    session_id: str | None = None,
    extra: str | None = None,
) -> Path:
    """
    Append a single log line to the active rotating log file.
    Rotates automatically when size >= MAX_FILE_SIZE.
    Returns the Path written to.
    """
    ts = datetime.utcnow().isoformat()
    p = get_active_log_file()
    line = (
        f"[{ts}] method={method} path={path} status={status} ip={ip} ua={ua} "
        f"session={session_id if session_id else '-'} extra={extra if extra else '-'}\n"
    )
    with open(p, "a", encoding="utf-8") as fh:
        fh.write(line)
    return p

# Per-session log helpers (optional XLSX)
def session_log_path(session_id: str, ext: str = "txt") -> Path:
    SESSION_DIR.mkdir(parents=True, exist_ok=True)
    return SESSION_DIR / f"session_{session_id}.{ext}"

def append_session_text(session_id: str, line: str):
    p = session_log_path(session_id, "txt")
    with open(p, "a", encoding="utf-8") as fh:
        fh.write(line + "\n")

def append_session_xlsx(session_id: str, row: List):
    if not OPENPYXL_AVAILABLE:
        # silently skip if not installed
        return
    p = session_log_path(session_id, "xlsx")
    if not p.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["timestamp", "method", "path", "status", "ip", "user_agent", "extra"])
        ws.append(row)
        wb.save(p)
    else:
        wb = load_workbook(p)
        ws = wb.active
        ws.append(row)
        wb.save(p)

# Compatibility wrappers (older names)
def append_text_log(session_id: str, line: str):
    return append_session_text(session_id, line)

def append_xlsx_log(session_id: str, row: List):
    return append_session_xlsx(session_id, row)
